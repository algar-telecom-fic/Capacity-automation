import pandas as pd
import re
import numpy as np
import openpyxl
import pkg_resources.py2_warn

def getData(dataFile):
    fo = open(dataFile)
    lines = fo.readlines()
    lengthFile = len(lines)
    body = []
    headers = []
    splitBody = []
    
    for i in range(lengthFile):
        #cria a variavel onde o programa deve começar a ler os dados (start)
        if(lines[i].startswith('License Usage')):
            start = i+2
        
        #cria a variavel onde o programa deve parar de ler os dados (end)
        elif(lines[i].startswith('(Number of')):
            h = (re.split(r'\s{2,}', lines[start]))       #splita o cabeçalho
            end = i - 1
            
            #Faz o split do corpo do arquivo
            for t in range(start+2, end+1):
                splitLine = (re.split(r'\s{2,}', lines[t][1:]))
                splitBody.append(splitLine)   #vetor de linhas splitadas
        
            headers.append(h)     #vetor de cabeçalhos (nesse caso temos apenas um cabeçalho)
            body.extend(splitBody)  #vetor final que guarda todo o arquivo splitado
            
            #trocando os espaços da coluna License Item por Underscore "_"
            # for t in range(len(body)):
            #     body[t][1] = body[t][1].replace(" ", "_")
            break;

    # Tratando a string vazia qua aparece no final (posição 7 do vetor interno)
    for t in body:
        del t[7]
        del t[6]

    for t in range(len(body)):
        for ti in range(len(body[t])):
            if(ti > 2 and (body[t][ti] != '' and body[t][ti] != '-')):
                body[t][ti] = float(body[t][ti])
			    

    return body


def createCsv(File):
    tuples = getData(File)
    df = pd.DataFrame(tuples, columns = ['License ID', 'License Item', 'Type', 'maximum_tuple_number', 'used_number', 'Usage-percent(%)'])
    return df


def renameColumns(df, columns): 
    newColumns = {}
    
    for c in columns:
        newColumns[c] = c.lower().replace(" ", "_")
    
    df = df.rename(columns=newColumns)

    return df

def generateCompleteReport(fileBase, createdFile, countDays, name_xlsx):
    created = createdFile
    created = renameColumns(created, list(created.columns))
    
    base = fileBase
    base = renameColumns(base, list(base.columns))

    completeReport = []
    
    for index,row in created.iterrows():
        rowBase = base.loc[base['license_id'] == row.license_id]
        if(row.maximum_tuple_number == 0):
            usage = 0
        else:
            usage = float(round(row.used_number/row.maximum_tuple_number, 4)*100)
            
        # monthlyGrowth = int(round(((row.used_number - rowBase.used_number.values[0])/countDays)))
        realGrowth = int(round(((row.used_number - rowBase.used_number.values[0])/countDays)*30))
        
        available = round(row.maximum_tuple_number - row.used_number)
        if realGrowth != 0:
            forecastNumber = round(available/realGrowth)  
        else:
            forecastNumber = 0
        
        if forecastNumber == 0: forecast = 'Estavel'
        elif forecastNumber >= 1 and forecastNumber <= 24: forecast = forecastNumber
        elif forecastNumber > 24: forecast = 'Maior que 2 Anos'
        elif forecastNumber < 0: forecast = 'Decrescimento'
            
        newRow = [row.license_id, row.license_item, row.type, int(row.maximum_tuple_number), int(row.used_number), usage, realGrowth, forecast]
        completeReport.append(newRow)
    
    report = pd.DataFrame(completeReport, columns=['ID', 'DESCRIPTION', 'TYPE', 'CAPACIDADE', 'UTILIZADO', 'UTILIZADO %','CRESCIMENTO MENSAL','PREVISAO DE ESGOTAMENTO / MES'])

    report.to_excel(name_xlsx, index=False)
    colorCells(name_xlsx)
    return report

def colorCells(name_xlsx):
	red_color = openpyxl.styles.PatternFill(start_color = 'FFFF0000', end_color = 'FFFF0000', fill_type = 'solid')
	yellow_color = openpyxl.styles.PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')
	green_color = openpyxl.styles.PatternFill(start_color = '008000', end_color = '008000', fill_type = 'solid')
	grey_color = openpyxl.styles.PatternFill(start_color = '4F4F4F', end_color = '4F4F4F', fill_type = 'solid')

	sheets_file = openpyxl.load_workbook(name_xlsx, data_only=True)
	sheet = sheets_file.worksheets[0]

	column = 6
	for row in range(2, 5000):
		if sheet.cell(column=column, row=row).value is None:
			break
		else:
			comp = float(sheet.cell(column=column, row=row).value)
			if comp >= 99.5:
				sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
				sheet.cell(column=column, row=row).fill = grey_color
			elif comp >= 90.5 and comp < 99.5:
				sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
				sheet.cell(column=column, row=row).fill = red_color
			elif comp >= 70.5 and comp < 90.5:
				sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
				sheet.cell(column=column, row=row).fill = yellow_color
			elif comp >= 50.5 and comp < 70.5:
				sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
				sheet.cell(column=column, row=row).fill = green_color
			else:
				sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'

	sheets_file.save(name_xlsx)

def run(current_file, previous_file, name_xlsx, days):
    current = createCsv(current_file)
    previous = createCsv(previous_file)
    generateCompleteReport(previous, current, days, name_xlsx)

if __name__ == '__main__':
	print('O formato dos arquivos finais é xlsx')
	input('Lembre de mudar os argumentos no txt, após ter feito isso tecle Enter: ')
	
	archive = open('arguments_licenses.txt','r')
	list_path = [lines[lines.find('=')+2:lines.find('\n')] for lines in archive if(lines != '\n')]
	archive.close()

	for argument in range(0, len(list_path), 4):
		run(list_path[argument], list_path[argument+1], list_path[argument+2] , int(list_path[argument+3]))

	input('Os arquivos estão prontos, tecle Enter para sair')