import pandas as pd
import re
import numpy as np
import openpyxl
import pkg_resources.py2_warn

def getTuples(fileName):
    f = open(fileName)
    lines = f.readlines()
    i = 0 
    start = end = -1
    maxLines = len(lines) - 1
    tuples = []
    headers = []
    
    # Separa as tuplas 
    while i <= maxLines: 
        if lines[i].startswith('Maximum number'):
            start = i + 2

        elif lines[i].startswith('(Number of '):
            h = (re.split(r'\s{2,}', lines[start]))
            end = i - 1
            tup = [t.split() for t in lines[start+2:end+1]]
            
            # Verifica se possui campo "Module number"
            if len(h) == 4:
                k = 0
                
                for t in tup:
                    t.insert(2,None)
                    
            tuples.extend(tup)
            headers.append(h)

        i = i + 1
    
    i = 0 
    maxTuples = len(tuples) - 1
    
    # Converte de string pra float quando for digito
    while i <= maxTuples:            
        j = 0
        
        while j <= len(tuples[i]) - 1:
            if j == 1:
                tuples[i][j] = int(tuples[i][j])
                
            elif tuples[i][j] is not None and tuples[i][j].isdigit():
                tuples[i][j] = float(tuples[i][j])
            j = j + 1
                
        i = i + 1
        
    return tuples

def defineTableType(fileName, premise):
    tables = {}
    if premise == 'SPO':
        df = pd.read_excel(fileName, sheet_name='TABELAS SPO')
    elif premise == 'ULA':  
        df = pd.read_excel(fileName, sheet_name='TABELAS ULA')
    elif premise == 'FAC':
        df = pd.read_excel(fileName, sheet_name='TABELAS FAC')
    
    # Separa as tabelas marcadas como únicas (não consideram intervalo de módulo)
    unique = df.OBSERVAÇÃO.str.contains('TABELA ÚNICA')
    result = df[unique]
    
    tableIDS = list(result['Table ID'])
    for tableID in tableIDS: tables[tableID] = None
    
    # Separa as tabelas marcadas como duplicada sem módulo (não considera intervalo de módulo)
    notUnique = ~df.OBSERVAÇÃO.str.contains('TABELA ÚNICA')
    noModule = df.OBSERVAÇÃO.str.contains('NÃO POSSUI MÓDULO')
    result = df[notUnique]
    result = result[noModule]
    
    tableIDS = list(result['Table ID'])
    for tableID in tableIDS: tables[tableID] = None
        
    # Separa as tabelas marcadas como duplicada e com intervalo de módulo a considerar
    notUnique = ~df.OBSERVAÇÃO.str.contains('TABELA ÚNICA')
    withModule = df.OBSERVAÇÃO.str.contains('COM MÓDULO')
    result = df[notUnique]
    result = result[withModule]
    result = result[['Table ID', 'OBSERVAÇÃO']]
    
    for index, row in result.iterrows():
        startIntervalo = row['OBSERVAÇÃO'].find('DO MÓDULO')
        endIntervalo = row['OBSERVAÇÃO'].find('.')
        intervalo = row['OBSERVAÇÃO']
        intervalo = intervalo[startIntervalo+len('DO MÓDULO'):endIntervalo]

        startIntervalo = intervalo.find('AO')
        start = int(intervalo[:startIntervalo])
        end = int(intervalo[startIntervalo+3:])
        intervalo = (start,end)

        tables[row['Table ID']] = intervalo
        
    return tables

def createFinalCSV(tuplesFile, tablesFile, premise):
    tuples = getTuples(tuplesFile)
    df = pd.DataFrame(tuples, 
                      columns=['Table_name', 'Table_ID', 'Module_number', 'Maximum_tuple_number', 'Used_Number'])
    tables = defineTableType(tablesFile, premise)
    notFound = {}
    treated = {}
    data = []
    
    for index, row in df.iterrows():
        try:
            tableType = tables[row['Table_ID']]
            
            if row['Table_ID'] in treated: continue
                
            elif tableType is None:
                treated[row['Table_ID']] = True
                data.append(row.tolist())
                
            else:
                # Realiza tratativa nas tabelas que consideram intervalo de módulo
                treated[row['Table_ID']] = True
                start = tableType[0]
                end = tableType[1]
                
                result = df[(df.Table_ID == row['Table_ID']) & ((df.Module_number >= start ) & (df.Module_number <= end))]
                maxTupleNumber = result['Maximum_tuple_number'].sum()
                usedNumber = result['Used_Number'].sum()
                d = [row['Table_name'], row['Table_ID'], start, maxTupleNumber, usedNumber]
                data.append(d)
            
        except KeyError:
            # Salva os IDs não tratados
            notFound[row['Table_ID']] = True
    
    treatedCSV = pd.DataFrame(data, 
                              columns=['Table_name', 'Table_ID', 'Module_number', 'Maximum_tuple_number', 'Used_Number'])
    treatedCSV = treatedCSV.sort_values(by=['Table_ID'])
    treatedCSV = treatedCSV[['Table_ID','Table_name', 'Module_number', 'Maximum_tuple_number', 'Used_Number']]
    
    return treatedCSV

def renameColumns(df, columns): 
    newColumns = {}
    
    for c in columns:
        newColumns[c] = c.lower().replace(" ", "_")
    
    df = df.rename(columns=newColumns)

    return df

def generateCompleteReport(fileBase, fileActual, countDays, name_xlsx):
    actual = fileActual
    base = fileBase
    base = renameColumns(base, list(base.columns))
    actual = renameColumns(actual, list(actual.columns))
    completeReport = []
    
    for index,row in actual.iterrows():
        rowBase = base.loc[base['table_id'] == row.table_id]
        usage = float(round(row.used_number/row.maximum_tuple_number, 4)*100)
        realGrowth = int(round(((row.used_number - rowBase.used_number.values[0])/countDays)*30))
        
        available = round(row.maximum_tuple_number - row.used_number)
        if realGrowth != 0:
            forecastNumber = round(available/realGrowth)
        else:
            forecastNumber = 0
        
        if forecastNumber == 0: forecast = 'Estavel'
        elif forecastNumber >= 1 and forecastNumber <= 24: forecast = forecastNumber
        elif forecastNumber > 24: forecast = 'Maior que 2 Anos'
        elif forecastNumber < 0: forecast = 'Decrescimento'
            
        newRow = [row.table_id, row.table_name, int(row.maximum_tuple_number), int(row.used_number), usage, realGrowth, forecast]
        completeReport.append(newRow)
    
    report = pd.DataFrame(completeReport, columns=['TABLE ID','TABLE NAME','CAPACIDADE','UTILIZADO','UTILIZADO %','CRESCIMENTO MENSAL','PREVISAO DE ESGOTAMENTO / MES'])

    report.to_excel(name_xlsx, index=False)
    colorCells(name_xlsx)
    return report

def colorCells(name_xlsx):
    red_color = openpyxl.styles.PatternFill(start_color = 'FFFF0000', end_color = 'FFFF0000', fill_type = 'solid')
    yellow_color = openpyxl.styles.PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')
    green_color = openpyxl.styles.PatternFill(start_color = '008000', end_color = '008000', fill_type = 'solid')
    grey_color = openpyxl.styles.PatternFill(start_color = '4F4F4F', end_color = '4F4F4F', fill_type = 'solid')

    sheets_file = openpyxl.load_workbook(name_xlsx, data_only=True)
    sheet = sheets_file.worksheets[0]

    column = 5
    for row in range(2, 5000):
        if sheet.cell(column=column, row=row).value is None:
            break
        else:
            comp = float(sheet.cell(column=column, row=row).value)
            if comp >= 99.5:
                sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
                sheet.cell(column=column, row=row).fill = grey_color
            elif comp >= 90.5 and comp < 99.5:
                sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
                sheet.cell(column=column, row=row).fill = red_color
            elif comp >= 70.5 and comp < 90.5:
                sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
                sheet.cell(column=column, row=row).fill = yellow_color
            elif comp >= 50.5 and comp < 70.5:
                sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'
                sheet.cell(column=column, row=row).fill = green_color
            else:
                sheet.cell(column=column, row=row).value = str(format(float(sheet.cell(column=column, row=row).value), '.2f')) + '%'

    sheets_file.save(name_xlsx)    

def run(current_file, previous_file, name_xlsx, days, softx):
    if softx == 0: premise = 'ULA'
    elif softx == 4: premise = 'SPO'
    elif softx == 8: premise = 'FAC'
    current = createFinalCSV(current_file, 'Premissas Tabelas.xlsx', premise)
    previous = createFinalCSV(previous_file,'Premissas Tabelas.xlsx', premise)
    generateCompleteReport(previous, current, days, name_xlsx)

if __name__ == '__main__':
    print('O formato dos arquivos finais é xlsx')
    input('Lembre de mudar os argumentos no txt, após ter feito isso tecle Enter: ')
    
    archive = open('arguments_tables.txt','r')
    list_path = [lines[lines.find('=')+2:lines.find('\n')] for lines in archive if(lines != '\n')]
    archive.close()

    for argument in range(0, len(list_path), 4):
        run(list_path[argument], list_path[argument+1], list_path[argument+2] , int(list_path[argument+3]), argument)

    input('Os arquivos estão prontos, tecle Enter para sair')